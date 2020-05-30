# -*- coding: utf-8 -*-
# Copyright (c) 2015, Bravo Logisitcs and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe, json
from frappe.model.document import Document
from frappe.utils import encode, cstr, cint, flt, comma_or
from frappe import _

import openpyxl
from io import StringIO
from openpyxl.styles import Font
from openpyxl import load_workbook
from frappe.utils.xlsxutils import handle_html
from csf_tz.after_sales_services.doctype.requested_payments.requested_payments import validate_requested_funds

class Export(Document):
	from csf_tz.after_sales_services.doctype.reference_payment_table.reference_payment_table import update_child_table
	
	def onload(self):
		if not self.company:
			self.company = frappe.defaults.get_user_default("Company") or frappe.defaults.get_global_default("company")
			
	def before_save(self):
		validate_requested_funds(self)


@frappe.whitelist(allow_guest=True)
def create_export(**args):
	args = frappe._dict(args)
	
	existing_export = frappe.db.get_value("Export", 
		{"booking_number": args.booking_number})
	
		
	if existing_export:
		doc = frappe.get_doc("Export", existing_export)
		doc.db_set("booking_received", args.documents_received_date)
		doc.db_set("client", args.customer)
		doc.db_set("export_type", args.export_type)
		doc.db_set("file_number", args.file_number)
		return existing_export
	else:
		new_export = frappe.new_doc("Export")
		new_export.update({
			"booking_number":args.booking_number,
			"booking_received": args.documents_received_date,
			"client": args.customer,
			"export_type": args.export_type,
			"file_number": args.file_number
		})
		new_export.insert(ignore_permissions=True)
		return new_export.name
		
@frappe.whitelist(allow_guest=True)
def download_ticts_loading_list(**args):
	args = frappe._dict(args)
	
	data = [['Test 5', 'Test 6', args.args], ['Test 7', 'Test 8']]
	
	sheet_name = "Loading List"
	wb = openpyxl.Workbook(write_only=True)
	ws = wb.create_sheet(sheet_name, 0)

	row1 = ws.row_dimensions[1]
	row1.font = Font(name='Calibri',bold=True)
	
	#data = [['Test 5', 'Test 6'], ['Test 7', 'Test 8']]
	
	for row in data:
		clean_row = []
		for item in row:
			if isinstance(item, basestring) and sheet_name != "Data Import Template":
				value = handle_html(item)
			else:
				value = item
			clean_row.append(value)

		ws.append(clean_row)

	xlsx_file = StringIO()
	wb.save(xlsx_file)
	
	frappe.response['filename'] = sheet_name + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'download'


@frappe.whitelist(allow_guest=True)
def check_export_status(**args):
	args = frappe._dict(args)
	#get export
	existing_export = frappe.db.get_value("Export",
	{"file_number": args.file_number})
	if existing_export:
		doc = frappe.get_doc("Export", existing_export)
		status=doc.status
		if status=='Open' or status=='Cancelled':
			frappe.msgprint("Cannot Close the File because it's EXPORT is not closed,Please Close the EXPORT first")
		else:
			return status
