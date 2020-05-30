# Copyright (c) 2013, Codes Consult and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe.utils import flt, getdate, cstr
from frappe import _
import ast
from datetime import datetime
from frappe.desk.reportview import export_query
from frappe.desk.query_report import run
from frappe.desk.query_report import get_columns_dict
from six import string_types
import os, json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from six import StringIO, string_types
from frappe.utils.xlsxutils import handle_html
import re

@frappe.whitelist()
def export_query():
	"""export from query reports"""

	data = frappe._dict(frappe.local.form_dict)

	del data["cmd"]
	if "csrf_token" in data:
		del data["csrf_token"]

	if isinstance(data.get("filters"), string_types):
		filters = json.loads(data["filters"])
	if isinstance(data.get("report_name"), string_types):
		report_name = data["report_name"]
	if isinstance(data.get("file_format_type"), string_types):
		file_format_type = data["file_format_type"]
	if isinstance(data.get("visible_idx"), string_types):
		visible_idx = json.loads(data.get("visible_idx"))
	else:
		visible_idx = None

	if file_format_type == "Excel":

		data = run(report_name, filters)
		data = frappe._dict(data)
		columns = get_columns_dict(data.columns)

		result = [[]]

		# add column headings
		for idx in range(len(data.columns)):
			result[0].append(columns[idx]["label"])

		# build table from dict
		if isinstance(data.result[0], dict):
			for i,row in enumerate(data.result):
				# only rows which are visible in the report
				if row and (i+1 in visible_idx):
					row_list = []
					for idx in range(len(data.columns)):
						row_list.append(row.get(columns[idx]["fieldname"],""))
					result.append(row_list)
				elif not row:
					result.append([])
		else:
			result = result + [d for i,d in enumerate(data.result) if (i+1 in visible_idx)]

		#from frappe.utils.xlsxutils import make_xlsx
		xlsx_file = make_xlsx(result, report_name)

		frappe.response['filename'] = report_name + '.xlsx'
		frappe.response['filecontent'] = xlsx_file.getvalue()
		frappe.response['type'] = 'binary'
		
def make_xlsx(data, report_name):
	wb = openpyxl.Workbook()
	ws = wb.create_sheet('All', 0)
	wb.guess_types = True
	customers = []
	
	ws.append([''])
	ws.append([''])
	ws.append([''])
	ws.append([''])
	ws.append([''])
	ws.append([''])
	ws.append([''])
	
	for row in data:
		clean_row = []
		if row[1] and row[1] not in customers and str(row[1]).upper() not in ['CUSTOMER', 'CLIENT']:
			customers.append(row[1])
		for item in row:
			if isinstance(item, string_types) and report_name != "Data Import Template":
				value = item
			else:
				value = item
			clean_row.append(value)

		ws.append(clean_row)
	ws = style_export(ws, report_name)
	
	del data[0][1]  #Remove the customer heading
	#Create a sheet for each customer
	if report_name not in ['Workshop Report', 'Workshop Daily Report']:
		for customer in customers:
			clean_title = customer.replace('\\', '')
			clean_title = clean_title.replace('/', '')
			wsc = wb.create_sheet(clean_title)
			wsc.append([''])
			wsc.append([''])
			wsc.append([''])
			wsc.append([''])
			wsc.append([''])
			wsc.append([''])
			wsc.append([''])
			wsc.append(data[0])
			for row in data:
				row = list(row)
				if isinstance(row[1], string_types) and handle_html(row[1]) == customer:
					clean_row = []
					del row[1]
					for item in row:
						if isinstance(item, string_types) and report_name != "Data Import Template":
							value = item
						else:
							value = item
						clean_row.append(value)
					wsc.append(clean_row)
			wsc = style_export(wsc, report_name)
	
	xlsx_file = StringIO()
	wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
	wb.save(xlsx_file)
	return xlsx_file
	
def style_export(sheet, report_name):
	logo = openpyxl.drawing.image.Image('assets/erpnext/images/logo.png', size=[110, 111])
	
	if report_name in ['Daily Customer Report - Imports', 'Daily Customer Report - Exports', 'Daily Customer Report - Transport', 'Border Clearance', 'Workshop Report', 'Imports Report',
						'Exports Report', 'Workshop Daily Report', 'Vehicle Status Report', 'Transport Assignment Report', 'Vehicles En Route to Border', 
						'Bond Report', 'Export Vehicles En Route to Warehouse']:
		#Set heading style
		alphabets = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
		for i in alphabets:
			cell1 = sheet[i + '8']
			cell2 = sheet['A'+i+'8']
			border_style = Border(left=Side(border_style='medium', color='FF000000'), right=Side(border_style='medium', color='FF000000'), top=Side(border_style='medium', color='FF000000'), bottom=Side(border_style='medium', color='FF000000'))
			if cell1.value:
				cell1.border = border_style
				cell1.font = Font(name='Calibri',size=9,bold=True)
				cell1.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True,shrink_to_fit=False,indent=0)
				
			if cell2.value:
				cell2.border = border_style
				cell2.font = Font(name='Calibri',size=9,bold=True)
				cell2.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True,shrink_to_fit=False,indent=0)
		
		sheet.add_image(logo, 'B2')
		sheet['I4'] = report_name + ' (' + datetime.now().strftime("%d-%m-%Y") + ')'
		row1 = sheet.row_dimensions[8]
		row1.height = 45
		
		#Set font for each cell in document
		for i in range(1, sheet.max_column+1):
			for j in range(1, sheet.max_row+1):
				sheet.cell(row = j, column = i).font = Font(name='Calibri',size=9)
		
	return sheet
		
	
	
	
